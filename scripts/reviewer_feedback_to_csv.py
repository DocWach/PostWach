"""Generate a flat CSV/XLSX adjudication view from a canonical reviewer_feedback.yaml.

The YAML stays the canonical source. The CSV is a derived adjudication grid; do not
edit the CSV directly. Re-run this script after any YAML change.

Usage:
    python reviewer_feedback_to_csv.py <path-to-reviewer_feedback.yaml> [--xlsx]
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    import yaml  # type: ignore
except ImportError:
    print("ERROR: PyYAML is required. Install with: pip install pyyaml", file=sys.stderr)
    sys.exit(1)


COLUMNS = [
    "comment_id",
    "reviewer_id",
    "reviewer_name",
    "date",
    "priority",
    "category",
    "agreement",
    "anchored_quote",
    "comment_text",
    "suggested_change",
    "effort_estimate",
    "cross_links",
    "paragraph_context",
    "adjudication_decision",
    "adjudication_notes",
    "implementation_status",
]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("yaml_path", help="Path to reviewer_feedback.yaml")
    ap.add_argument("--xlsx", action="store_true", help="Also emit .xlsx alongside the .csv")
    args = ap.parse_args()

    yaml_path = Path(args.yaml_path)
    if not yaml_path.exists():
        print(f"ERROR: {yaml_path} not found", file=sys.stderr)
        sys.exit(1)

    data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
    reviewer_lookup = {r["id"]: r["name"] for r in data.get("reviewers", [])}

    rows = []
    for c in data.get("comments", []):
        ia = c.get("independent_assessment", {}) or {}
        adj = c.get("adjudication", {}) or {}
        rows.append({
            "comment_id": c.get("id", ""),
            "reviewer_id": c.get("reviewer_id", ""),
            "reviewer_name": reviewer_lookup.get(c.get("reviewer_id", ""), ""),
            "date": c.get("date", ""),
            "priority": ia.get("priority", ""),
            "category": ia.get("category", ""),
            "agreement": ia.get("agreement", ""),
            "anchored_quote": (c.get("anchored_quote") or "").strip(),
            "comment_text": (c.get("text") or "").strip(),
            "suggested_change": (ia.get("suggested_change") or "").strip(),
            "effort_estimate": ia.get("effort_estimate", ""),
            "cross_links": ", ".join(ia.get("cross_links", []) or []),
            "paragraph_context": (c.get("paragraph_context") or "").strip(),
            "adjudication_decision": adj.get("decision", ""),
            "adjudication_notes": (adj.get("notes") or "").strip(),
            "implementation_status": adj.get("implementation_status", ""),
        })

    csv_path = yaml_path.with_suffix(".csv")
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"Wrote {csv_path} ({len(rows)} rows)")

    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("WARN: openpyxl not installed; skipping .xlsx. Install with: pip install openpyxl", file=sys.stderr)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "reviewer_feedback"
        ws.append(COLUMNS)
        for row in rows:
            ws.append([row[c] for c in COLUMNS])

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        for col_idx, _ in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", horizontal="left")

        widths = {
            "comment_id": 10,
            "reviewer_id": 10,
            "reviewer_name": 22,
            "date": 22,
            "priority": 10,
            "category": 14,
            "agreement": 16,
            "anchored_quote": 38,
            "comment_text": 60,
            "suggested_change": 60,
            "effort_estimate": 16,
            "cross_links": 14,
            "paragraph_context": 60,
            "adjudication_decision": 22,
            "adjudication_notes": 40,
            "implementation_status": 22,
        }
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)

        for row_idx in range(2, ws.max_row + 1):
            for col_idx, _ in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row_idx].height = 95

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        xlsx_path = yaml_path.with_suffix(".xlsx")
        wb.save(xlsx_path)
        print(f"Wrote {xlsx_path}")


if __name__ == "__main__":
    main()
