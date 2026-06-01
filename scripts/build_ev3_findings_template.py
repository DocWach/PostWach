"""Build the EV3 model validation findings spreadsheet template.

Output: 01 PostWach/docs/EV3_Model_Validation_Findings_Template_2026-04-29.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="1F4E78", size=11)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(italic=True, color="595959", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP
        cell.border = BORDER


def style_data_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = WRAP_TOP
        cell.border = BORDER


def set_column_widths(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def add_dropdown(ws, options: list[str], col_letter: str, row_start: int, row_end: int) -> None:
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col_letter}{row_start}:{col_letter}{row_end}")
    ws.add_data_validation(dv)


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    set_column_widths(ws, {1: 22, 2: 90})

    ws["A1"] = "EV3 Model Validation: Findings Workbook"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:B1")

    ws["A2"] = "Companion to:"
    ws["B2"] = "EV3_Model_Validation_Student_Brief_2026-04-29.pdf"
    ws["A3"] = "Repo:"
    ws["B3"] = "https://github.com/DocWach/Lego-EV3-Mindstorm-Models"

    ws["A5"] = "How to use"
    ws["A5"].font = SUBHEADER_FONT

    instructions = [
        ("1.", "Read the student brief PDF first. This workbook holds the structured findings; the brief explains the workflow."),
        ("2.", "Phase 1 (BOM): fill in 'Phase 1 BOM' as you parse the three kit SysML files."),
        ("3.", "Phase 2 (Walk-through): use 'Findings Log' for every smell. Aim for at least 12 substantive entries."),
        ("4.", "Phase 3 studies: use the three dedicated sheets ('Phase 3A Multiplicity', 'Phase 3B Interfaces', 'Phase 3C Attributes')."),
        ("5.", "Verdict: complete 'Verdict Summary' last. Pick extend, refactor, or rebuild and justify with your strongest findings."),
        ("6.", "Cross-reference: cite F-NN IDs from the Findings Log in your Word report."),
    ]
    for i, (n, t) in enumerate(instructions, start=6):
        ws.cell(row=i, column=1, value=n).alignment = WRAP_TOP
        ws.cell(row=i, column=2, value=t).alignment = WRAP_TOP

    ws["A13"] = "Severity definitions"
    ws["A13"].font = SUBHEADER_FONT
    sev = [
        ("critical", "Model would mislead a downstream user (PLMr, SysMLv2 hive, future student). Wrong part identity, wrong kit composition, mating constraint that contradicts physics."),
        ("major", "Structural choice that should change. Duplicate definitions, multiplicity captured inconsistently, attribute missing where a real-world specification exists."),
        ("minor", "Cosmetic or local issue. Naming inconsistency, comment missing, attribute formatting. Worth logging but not blocking."),
    ]
    for i, (s, d) in enumerate(sev, start=14):
        ws.cell(row=i, column=1, value=s).font = Font(bold=True)
        ws.cell(row=i, column=1).alignment = WRAP_TOP
        ws.cell(row=i, column=2, value=d).alignment = WRAP_TOP

    ws["A18"] = "Phase codes"
    ws["A18"].font = SUBHEADER_FONT
    phases = [
        ("1", "Hands-on orientation + kit BOM piece-count validation."),
        ("2", "Inductive walk-through smell log."),
        ("3A", "Multiplicity and roll-up study."),
        ("3B", "Interface capture study."),
        ("3C", "Attribute correctness with physical measurement."),
        ("ad-hoc", "Surprise findings outside any defined phase."),
    ]
    for i, (p, d) in enumerate(phases, start=19):
        ws.cell(row=i, column=1, value=p).font = Font(bold=True)
        ws.cell(row=i, column=1).alignment = WRAP_TOP
        ws.cell(row=i, column=2, value=d).alignment = WRAP_TOP

    ws["A26"] = "Reviewer notes"
    ws["A26"].font = SUBHEADER_FONT
    ws["B26"] = "Leave the 'PostWach Notes' columns blank. PostWach fills these during review."
    ws["B26"].fill = NOTE_FILL
    ws["B26"].font = NOTE_FONT
    ws["B26"].alignment = WRAP_TOP


def build_findings_log(wb: Workbook) -> None:
    ws = wb.create_sheet("Findings Log")
    set_column_widths(ws, {1: 8, 2: 9, 3: 30, 4: 32, 5: 11, 6: 45, 7: 35, 8: 35, 9: 14, 10: 30})

    headers = [
        "ID", "Phase", "Title", "File (sysml/...)", "Severity",
        "Observation", "Why it matters", "Proposed fix or question",
        "Status", "PostWach Notes",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    seed = [
        "F-01", "2",
        "Duplicate brick / motors / sensors definitions across electronics/ and hardware/",
        "sysml/electronics/EV3Brick.sysml; sysml/hardware/EV3Brick.sysml",
        "major",
        "Three file pairs (EV3Brick.sysml, EV3Motors.sysml, EV3Sensors.sysml) appear in both electronics/ and hardware/. Determine whether this is intentional layering (parts vs system-level specs) or unresolved duplication.",
        "If unresolved duplication, two definitions can drift; downstream users do not know which is authoritative.",
        "Document the difference between the two copies; recommend whether to unify, retain with explicit role, or delete one.",
        "open", "",
    ]
    ws.append(seed)
    style_data_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 75

    blank_rows = 60
    for i in range(blank_rows):
        row_num = 3 + i
        ws.cell(row=row_num, column=1, value=f"F-{i + 2:02d}")
        style_data_row(ws, row_num, len(headers))
        ws.row_dimensions[row_num].height = 60

    add_dropdown(ws, ["1", "2", "3A", "3B", "3C", "ad-hoc"], "B", 2, 2 + blank_rows)
    add_dropdown(ws, ["critical", "major", "minor"], "E", 2, 2 + blank_rows)
    add_dropdown(ws, ["open", "needs-review", "resolved", "wont-fix"], "I", 2, 2 + blank_rows)

    ws.freeze_panes = "A2"


def build_phase1_bom(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase 1 BOM")
    set_column_widths(ws, {1: 14, 2: 12, 3: 28, 4: 12, 5: 14, 6: 11, 7: 30})

    ws["A1"] = "Phase 1 Kit BOM Piece-Count Validation"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:G1")

    ws["A2"] = "Three kits to validate. Two checks per kit: internal sum (does the SysML line items sum to claimed total?) and external comparison (do the line items match the published BrickLink or Brickset inventory?). Add one row per part. Use the 'Kit' column to filter."
    ws["A2"].alignment = WRAP_TOP
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 50

    ws["A4"] = "Summary"
    ws["A4"].font = SUBHEADER_FONT

    summary_headers = ["Kit", "Set name", "Claimed total (SysML)", "Sum of SysML lines (formula)", "Sum of published (formula)", "Internal match?", "External match?"]
    for c, h in enumerate(summary_headers, start=1):
        ws.cell(row=5, column=c, value=h)
    style_header(ws, 5, len(summary_headers))
    ws.row_dimensions[5].height = 30

    summary_rows = [
        (6, "31313", "Home Edition", 601),
        (7, "45544", "Education Core", 541),
        (8, "45560", "Expansion", 853),
    ]
    for row, kit, name, total in summary_rows:
        ws.cell(row=row, column=1, value=kit)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=total)
        ws.cell(row=row, column=4, value=f'=SUMIF(BOM_Kit,"{kit}",BOM_SysML_Qty)')
        ws.cell(row=row, column=5, value=f'=SUMIF(BOM_Kit,"{kit}",BOM_Pub_Qty)')
        ws.cell(row=row, column=6, value=f"=IF(C{row}=D{row},\"yes\",\"no\")")
        ws.cell(row=row, column=7, value=f"=IF(D{row}=E{row},\"yes\",\"no\")")
        style_data_row(ws, row, len(summary_headers))

    ws["A10"] = "Line-item table"
    ws["A10"].font = SUBHEADER_FONT

    line_headers = ["Kit", "LDraw ID", "Part Name", "SysML Qty", "Published Qty", "Match?", "Notes"]
    for c, h in enumerate(line_headers, start=1):
        ws.cell(row=11, column=c, value=h)
    style_header(ws, 11, len(line_headers))
    ws.row_dimensions[11].height = 30

    blank_rows = 400
    for i in range(blank_rows):
        row_num = 12 + i
        ws.cell(row=row_num, column=6, value=f'=IF(AND(ISNUMBER(D{row_num}),ISNUMBER(E{row_num})),IF(D{row_num}=E{row_num},"yes","no"),"")')
        style_data_row(ws, row_num, len(line_headers))

    add_dropdown(ws, ["31313", "45544", "45560"], "A", 12, 11 + blank_rows)

    _add_named_range(wb, "BOM_Kit", ws.title, 1, 12, 11 + blank_rows)
    _add_named_range(wb, "BOM_SysML_Qty", ws.title, 4, 12, 11 + blank_rows)
    _add_named_range(wb, "BOM_Pub_Qty", ws.title, 5, 12, 11 + blank_rows)

    ws.freeze_panes = "A12"


def _add_named_range(wb: Workbook, name: str, sheet: str, col: int, row_start: int, row_end: int) -> None:
    from openpyxl.workbook.defined_name import DefinedName
    col_letter = get_column_letter(col)
    safe_sheet = f"'{sheet}'"
    ref = f"{safe_sheet}!${col_letter}${row_start}:${col_letter}${row_end}"
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def build_phase3a(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase 3A Multiplicity")
    set_column_widths(ws, {1: 6, 2: 12, 3: 28, 4: 30, 5: 35, 6: 35, 7: 35, 8: 12})

    ws["A1"] = "Phase 3A: Multiplicity and Roll-Up"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:H1")

    ws["A2"] = "Pick 5 parts that appear in multiple kits at varied quantities. For each, document how the model captures multiplicity now and what the rule should be."
    ws["A2"].alignment = WRAP_TOP
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 36

    headers = ["#", "LDraw ID", "Part Name", "SysML File(s)", "How multiplicity is captured now", "Where you think it should be captured", "Proposed rule for the library", "Severity"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(headers))
    ws.row_dimensions[4].height = 30

    for i in range(8):
        row_num = 5 + i
        ws.cell(row=row_num, column=1, value=i + 1 if i < 5 else "")
        style_data_row(ws, row_num, len(headers))
        ws.row_dimensions[row_num].height = 70

    add_dropdown(ws, ["critical", "major", "minor"], "H", 5, 12)
    ws.freeze_panes = "A5"


def build_phase3b(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase 3B Interfaces")
    set_column_widths(ws, {1: 6, 2: 25, 3: 30, 4: 40, 5: 40, 6: 18, 7: 30, 8: 12})

    ws["A1"] = "Phase 3B: Interface Capture"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:H1")

    ws["A2"] = "Pick 5 connection scenarios. Use the physical kit to verify mating. Document what the model says, what the kit does, and whether the model represents the constraint."
    ws["A2"].alignment = WRAP_TOP
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 36

    headers = ["#", "Connection type", "Parts involved", "Physical behavior (from kit)", "What the SysML model says", "Constraint represented?", "Notes / proposed fix", "Severity"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(headers))
    ws.row_dimensions[4].height = 30

    seeds = [
        (1, "pin-into-beam", "friction pin (LDraw 2780) + Technic beam"),
        (2, "axle-into-gear", "axle + gear with cross-axle hole"),
        (3, "gear-into-gear", "two spur gears (e.g., 24T + 40T)"),
        (4, "motor-shaft-into-coupler", "Large servo motor + axle/coupler"),
        (5, "cable-into-port", "EV3 cable + brick port or motor port"),
    ]
    for i, (n, ctype, parts) in enumerate(seeds, start=5):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=ctype)
        ws.cell(row=i, column=3, value=parts)
        style_data_row(ws, i, len(headers))
        ws.row_dimensions[i].height = 80

    add_dropdown(ws, ["yes", "partial", "no", "n/a"], "F", 5, 9)
    add_dropdown(ws, ["critical", "major", "minor"], "H", 5, 9)
    ws.freeze_panes = "A5"


def build_phase3c(wb: Workbook) -> None:
    ws = wb.create_sheet("Phase 3C Attributes")
    set_column_widths(ws, {1: 6, 2: 12, 3: 28, 4: 22, 5: 18, 6: 18, 7: 22, 8: 10, 9: 30})

    ws["A1"] = "Phase 3C: Attribute Correctness with Physical Measurement"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:I1")

    ws["A2"] = "Pick 8 parts where attributes are measurable from the physical kit. Compare the model attribute against the measured value. Flag any attribute missing where a real-world specification exists."
    ws["A2"].alignment = WRAP_TOP
    ws["A2"].fill = NOTE_FILL
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 36

    headers = ["#", "LDraw ID", "Part Name", "Attribute tested", "SysML value", "Measured value", "Method (count holes, protractor, ruler, datasheet)", "Match?", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header(ws, 4, len(headers))
    ws.row_dimensions[4].height = 30

    seeds = [
        (1, "", "Beam (any straight beam)", "length in modules", "", "", "count holes x 8 mm"),
        (2, "", "Bracket (any angular)", "angle", "", "", "protractor or known-angle reference"),
        (3, "", "Axle", "length in studs", "", "", "count studs"),
        (4, "", "Spur gear", "tooth count", "", "", "count teeth"),
        (5, "", "Cable", "length", "", "", "ruler"),
        (6, "", "EV3 large servo motor", "rated speed or stall torque", "", "", "LEGO datasheet"),
        (7, "", "Color sensor", "wavelength range or modes", "", "", "LEGO datasheet"),
        (8, "", "Tire", "outer diameter", "", "", "ruler"),
    ]
    for i, (n, ldraw, name, attr, sysml, measured, method) in enumerate(seeds, start=5):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=ldraw)
        ws.cell(row=i, column=3, value=name)
        ws.cell(row=i, column=4, value=attr)
        ws.cell(row=i, column=5, value=sysml)
        ws.cell(row=i, column=6, value=measured)
        ws.cell(row=i, column=7, value=method)
        style_data_row(ws, i, len(headers))
        ws.row_dimensions[i].height = 50

    add_dropdown(ws, ["yes", "close", "no", "missing-attribute"], "H", 5, 12)
    ws.freeze_panes = "A5"


def build_verdict(wb: Workbook) -> None:
    ws = wb.create_sheet("Verdict Summary")
    set_column_widths(ws, {1: 22, 2: 60, 3: 50})

    ws["A1"] = "Verdict Summary"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:C1")

    ws["A3"] = "Verdict"
    ws["A3"].font = SUBHEADER_FONT
    ws["B3"] = ""
    ws["B3"].alignment = CENTER
    ws["B3"].font = Font(bold=True, size=12)
    add_dropdown(ws, ["extend", "refactor", "rebuild"], "B", 3, 3)

    ws["A4"] = "Verdict reasoning"
    ws["A4"].font = SUBHEADER_FONT
    ws["B4"] = ""
    ws["B4"].alignment = WRAP_TOP
    ws.row_dimensions[4].height = 80

    ws["A6"] = "Top findings supporting the verdict"
    ws["A6"].font = SUBHEADER_FONT

    headers = ["Rank", "Finding ID (F-NN)", "Why this finding drives the verdict"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=7, column=c, value=h)
    style_header(ws, 7, len(headers))
    ws.row_dimensions[7].height = 30

    for i in range(5):
        row_num = 8 + i
        ws.cell(row=row_num, column=1, value=i + 1)
        style_data_row(ws, row_num, len(headers))
        ws.row_dimensions[row_num].height = 60

    ws["A14"] = "Open questions for PostWach"
    ws["A14"].font = SUBHEADER_FONT
    for i in range(5):
        row_num = 15 + i
        ws.cell(row=row_num, column=1, value=f"Q{i + 1}")
        style_data_row(ws, row_num, 3)
        ws.row_dimensions[row_num].height = 50

    ws["A21"] = "Method notes"
    ws["A21"].font = SUBHEADER_FONT
    ws["B21"] = "Tools used, scripts written, total hours spent."
    ws["B21"].alignment = WRAP_TOP
    ws["B21"].fill = NOTE_FILL
    ws["B21"].font = NOTE_FONT
    ws["B22"] = ""
    ws.row_dimensions[22].height = 100


def main() -> None:
    out_dir = Path(r"C:\Users\pfwac\OneDrive - University of Arizona\Documents\03 Projects\00 My Research\01 PostWach\docs")
    out_path = out_dir / "EV3_Model_Validation_Findings_Template_2026-04-29.xlsx"

    wb = Workbook()
    build_readme(wb)
    build_findings_log(wb)
    build_phase1_bom(wb)
    build_phase3a(wb)
    build_phase3b(wb)
    build_phase3c(wb)
    build_verdict(wb)

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
